import os
import django
from openpyxl import load_workbook


# Set up Django environment
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'crs.settings')
django.setup()

# after loading django, then we import model
from django.contrib.auth import get_user_model
User=get_user_model()

# Load the Excel file
workbook = load_workbook('fake12professors.xlsx')  # Provide the path to your Excel file
sheet = workbook.active

# Assuming the structure of the Excel file is: first_name, last_name, email, password
for row in sheet.iter_rows(min_row=2, values_only=True):
    first_name, last_name, email, password = row
    
    user = User.objects.create_user(email=email, password=password)
    user.first_name = first_name
    user.last_name = last_name
    user.role='Professor'
    
    user.save()
    
    print(f"User {email} created successfully.")
