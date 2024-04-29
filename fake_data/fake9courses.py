import os
import django
from openpyxl import load_workbook


# Set up Django environment
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'crs.settings')
django.setup()

# after loading django, then we import model
from registrar.models import AllCourses

# Get the directory of the current script
current_directory = os.path.dirname(os.path.abspath(__file__))

# Construct the full path to the Excel file
excel_file_path = os.path.join(current_directory, 'fake_data', 'fake9courses.xlsx')


# Load the Excel file
workbook = load_workbook('fake9courses.xlsx')  # Provide the path to your Excel file
sheet = workbook.active

# Assuming the structure of the Excel file is: first_name, last_name, email, password
for row in sheet.iter_rows(min_row=2, values_only=True):
    code, name = row
    
    user = AllCourses(code=code, name=name)
    user.save()
    
    print(f"Course {code} - {name} created successfully.")
